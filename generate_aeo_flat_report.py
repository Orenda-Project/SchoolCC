"""
AEO Schools Flat Report Generator
Queries all AEO users and their assigned schools from the production database,
resolves EMIS numbers to canonical school names, and generates a formatted
single-table Excel report with one row per AEO.

Output columns:
  # | AEO Name | Status | Markaz | School Names (newline-separated) | EMIS Numbers (newline-separated)
"""

import json
import os
import re
import sys
from datetime import datetime

import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

def get_database_url():
    """Read DATABASE_URL from .env file."""
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
    with open(env_path, 'r') as f:
        for line in f:
            line = line.strip()
            if line.startswith('DATABASE_URL='):
                return line[len('DATABASE_URL='):]
    raise ValueError("DATABASE_URL not found in .env file")


def parse_school_reference(ref_str):
    """
    Parse a school reference string from assigned_schools.
    Format: "School Name (EMIS_Number)" -> (school_name, emis)
    Or just: "School Name"              -> (school_name, None)
    """
    ref_str = ref_str.strip()
    match = re.match(r'^(.+?)\s*\(([A-Za-z0-9\-]+)\)\s*$', ref_str)
    if match:
        return match.group(1).strip(), match.group(2).strip()
    return ref_str, None


# ---------------------------------------------------------------------------
# Data query
# ---------------------------------------------------------------------------

def query_aeo_data(conn):
    """
    Query all AEO users and resolve their assigned_schools to canonical
    school names and EMIS numbers from the schools table.
    Returns a sorted list of dicts.
    """
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # 1. Fetch all AEO users
    cur.execute("""
        SELECT id, name, status, markaz_name, assigned_schools
        FROM users
        WHERE role = 'AEO'
        ORDER BY name ASC;
    """)
    aeos = cur.fetchall()
    print(f"Found {len(aeos)} AEO users")

    # 2. Fetch all schools for lookup
    cur.execute("""
        SELECT id, name, emis_number
        FROM schools
        ORDER BY name ASC;
    """)
    all_schools = cur.fetchall()
    print(f"Total schools in database: {len(all_schools)}")

    # Build lookup maps
    emis_to_school = {}
    name_to_school = {}
    name_lower_to_school = {}
    for school in all_schools:
        emis_to_school[school['emis_number']] = {
            'name': school['name'],
            'emis_number': school['emis_number'],
        }
        name_to_school[school['name']] = {
            'name': school['name'],
            'emis_number': school['emis_number'],
        }
        name_lower_to_school[school['name'].lower().strip()] = {
            'name': school['name'],
            'emis_number': school['emis_number'],
        }

    # 3. Resolve each AEO's assigned schools
    result = []
    for aeo in aeos:
        assigned = aeo.get('assigned_schools')
        if isinstance(assigned, str):
            try:
                assigned = json.loads(assigned)
            except json.JSONDecodeError:
                assigned = []
        if not isinstance(assigned, list):
            assigned = []

        schools_list = []
        for ref in assigned:
            if not ref:
                continue
            ref = str(ref)
            parsed_name, emis = parse_school_reference(ref)
            school_info = None

            # Strategy 1: Match by EMIS number (most reliable)
            if emis and emis in emis_to_school:
                school_info = emis_to_school[emis]
            # Strategy 2: Match by exact name
            elif parsed_name in name_to_school:
                school_info = name_to_school[parsed_name]
            # Strategy 3: Match by case-insensitive name
            elif parsed_name.lower().strip() in name_lower_to_school:
                school_info = name_lower_to_school[parsed_name.lower().strip()]
            # Strategy 4: Fallback - use the parsed reference as-is
            else:
                school_info = {
                    'name': parsed_name,
                    'emis_number': emis if emis else 'N/A',
                }

            schools_list.append(school_info)

        # Sort schools by name within each AEO
        schools_list.sort(key=lambda s: s['name'])

        result.append({
            'name': aeo.get('name', 'N/A'),
            'status': aeo.get('status', 'N/A'),
            'markaz': aeo.get('markaz_name', 'N/A') or 'N/A',
            'schools': schools_list,
        })

    cur.close()
    return result


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def create_excel_report(aeo_data, output_path):
    """
    Create a well-formatted Excel report with one row per AEO.
    Columns: # | AEO Name | Status | Markaz | School Names | EMIS Numbers
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "AEO Schools Report"

    # -- Styles --
    header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True
    )
    header_border = Border(
        left=Side(style='thin', color='1F3864'),
        right=Side(style='thin', color='1F3864'),
        top=Side(style='thin', color='1F3864'),
        bottom=Side(style='thin', color='1F3864'),
    )

    data_font = Font(name='Calibri', size=11)
    data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)
    data_alignment_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='top', wrap_text=False)

    thin_border = Border(
        left=Side(style='thin', color='B4C6E7'),
        right=Side(style='thin', color='B4C6E7'),
        top=Side(style='thin', color='B4C6E7'),
        bottom=Side(style='thin', color='B4C6E7'),
    )

    alt_row_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

    # -- Column widths --
    col_widths = {
        'A': 5,    # #
        'B': 30,   # AEO Name
        'C': 10,   # Status
        'D': 20,   # Markaz
        'E': 50,   # School Names
        'F': 15,   # EMIS Numbers
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # -- Header row --
    headers = ['#', 'AEO Name', 'Status', 'Markaz', 'School Names', 'EMIS Numbers']
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
    ws.row_dimensions[1].height = 25

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # -- Data rows --
    for row_idx, aeo in enumerate(aeo_data, start=1):
        excel_row = row_idx + 1  # +1 because row 1 is the header

        # Build newline-separated strings for schools and EMIS
        school_names = '\n'.join(s['name'] for s in aeo['schools']) if aeo['schools'] else 'No schools assigned'
        emis_numbers = '\n'.join(str(s['emis_number']) for s in aeo['schools']) if aeo['schools'] else ''

        # Column A: row number
        cell_num = ws.cell(row=excel_row, column=1, value=row_idx)
        cell_num.font = data_font
        cell_num.alignment = center_alignment
        cell_num.border = thin_border

        # Column B: AEO Name
        cell_name = ws.cell(row=excel_row, column=2, value=aeo['name'])
        cell_name.font = data_font
        cell_name.alignment = data_alignment
        cell_name.border = thin_border

        # Column C: Status
        cell_status = ws.cell(row=excel_row, column=3, value=aeo['status'])
        cell_status.font = data_font
        cell_status.alignment = data_alignment
        cell_status.border = thin_border

        # Column D: Markaz
        cell_markaz = ws.cell(row=excel_row, column=4, value=aeo['markaz'])
        cell_markaz.font = data_font
        cell_markaz.alignment = data_alignment
        cell_markaz.border = thin_border

        # Column E: School Names (wrap text for multi-line)
        cell_schools = ws.cell(row=excel_row, column=5, value=school_names)
        cell_schools.font = data_font
        cell_schools.alignment = data_alignment_wrap
        cell_schools.border = thin_border

        # Column F: EMIS Numbers (wrap text for multi-line)
        cell_emis = ws.cell(row=excel_row, column=6, value=emis_numbers)
        cell_emis.font = data_font
        cell_emis.alignment = data_alignment_wrap
        cell_emis.border = thin_border

        # Alternate row shading
        if row_idx % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=excel_row, column=col).fill = alt_row_fill

        # Auto-adjust row height based on the number of schools
        num_schools = len(aeo['schools']) if aeo['schools'] else 1
        # ~15 points per line of text, minimum 20
        row_height = max(20, num_schools * 15)
        ws.row_dimensions[excel_row].height = row_height

    # -- Auto-filter --
    ws.auto_filter.ref = f"A1:F{len(aeo_data) + 1}"

    # -- Save --
    wb.save(output_path)
    print(f"\nExcel report saved to: {output_path}")
    print(f"Total AEOs: {len(aeo_data)}")
    total_schools = sum(len(a['schools']) for a in aeo_data)
    print(f"Total school assignments: {total_schools}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    db_url = get_database_url()
    print("Connecting to database...")

    conn = psycopg2.connect(db_url)
    conn.set_session(readonly=True)
    print("Connected (read-only mode)\n")

    try:
        aeo_data = query_aeo_data(conn)

        output_path = '/home/taleemabad/taleemabad/SchoolCC/AEO_Schools_Report.xlsx'
        create_excel_report(aeo_data, output_path)

        # Console summary
        print("\n--- AEO Summary ---")
        for i, aeo in enumerate(aeo_data, 1):
            num_schools = len(aeo['schools'])
            print(f"  {i:2d}. {aeo['name']:<30s} Status: {aeo['status']:<10s} Markaz: {aeo['markaz']:<20s} Schools: {num_schools}")
    finally:
        conn.close()
        print("\nDatabase connection closed.")


if __name__ == '__main__':
    main()
