"""
AEO Schools Report Generator
Queries all AEO users and their assigned schools from the production database,
then generates a formatted Excel report.
"""

import json
import os
import re
from datetime import datetime

import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def get_database_url():
    """Read DATABASE_URL from .env file."""
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
    with open(env_path, 'r') as f:
        for line in f:
            line = line.strip()
            if line.startswith('DATABASE_URL='):
                return line[len('DATABASE_URL='):]
    raise ValueError("DATABASE_URL not found in .env file")


def parse_school_reference(ref_str):
    """
    Parse a school reference string from assigned_schools.
    Formats found:
      - "School Name (EMIS_Number)" e.g. "GPS KALRI (37330349)"
      - "School Name" without EMIS e.g. "GGPS CARRIAGE FACTORY"
    Returns (school_name, emis_or_none).
    """
    ref_str = ref_str.strip()
    # Match pattern: "Name (EMIS)" where EMIS can contain digits and hyphens
    match = re.match(r'^(.+?)\s*\(([A-Za-z0-9\-]+)\)\s*$', ref_str)
    if match:
        return match.group(1).strip(), match.group(2).strip()
    return ref_str, None


def query_aeo_data(conn):
    """Query all AEO users with their assigned schools resolved to school names and EMIS numbers."""
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Step 1: Get all AEO users
    cur.execute("""
        SELECT
            id,
            name,
            phone_number,
            markaz_name,
            assigned_schools
        FROM users
        WHERE role = 'AEO'
        ORDER BY name ASC;
    """)
    aeos = cur.fetchall()
    print(f"Found {len(aeos)} AEO users")

    # Step 2: Parse all school references from assigned_schools
    # The assigned_schools field contains strings like "School Name (EMIS)" or just "School Name"
    all_emis_numbers = set()
    all_school_names = set()
    for aeo in aeos:
        assigned = aeo.get('assigned_schools')
        if assigned:
            if isinstance(assigned, str):
                assigned = json.loads(assigned)
            if isinstance(assigned, list):
                for ref in assigned:
                    if ref:
                        name, emis = parse_school_reference(str(ref))
                        if emis:
                            all_emis_numbers.add(emis)
                        all_school_names.add(name)

    print(f"Total unique EMIS numbers found: {len(all_emis_numbers)}")
    print(f"Total unique school names found: {len(all_school_names)}")

    # Step 3: Load all schools from the database for lookup
    cur.execute("""
        SELECT id, name, emis_number
        FROM schools
        ORDER BY name ASC;
    """)
    all_schools = cur.fetchall()
    print(f"Total schools in database: {len(all_schools)}")

    # Build lookup maps
    emis_to_school = {}
    name_to_school = {}
    name_lower_to_school = {}
    for school in all_schools:
        emis_to_school[school['emis_number']] = {
            'name': school['name'],
            'emis_number': school['emis_number']
        }
        name_to_school[school['name']] = {
            'name': school['name'],
            'emis_number': school['emis_number']
        }
        name_lower_to_school[school['name'].lower().strip()] = {
            'name': school['name'],
            'emis_number': school['emis_number']
        }

    # Step 4: Build the result structure
    resolved_count = 0
    unresolved_count = 0
    result = []
    for aeo in aeos:
        assigned = aeo.get('assigned_schools')
        if isinstance(assigned, str):
            assigned = json.loads(assigned)
        if not isinstance(assigned, list):
            assigned = []

        schools_list = []
        for ref in assigned:
            if not ref:
                continue
            ref = str(ref)
            name, emis = parse_school_reference(ref)
            school_info = None

            # Strategy 1: Match by EMIS number (most reliable)
            if emis and emis in emis_to_school:
                school_info = emis_to_school[emis]
            # Strategy 2: Match by exact name
            elif name in name_to_school:
                school_info = name_to_school[name]
            # Strategy 3: Match by case-insensitive name
            elif name.lower().strip() in name_lower_to_school:
                school_info = name_lower_to_school[name.lower().strip()]
            # Strategy 4: Use the parsed name and EMIS as-is from the reference string
            else:
                school_info = {
                    'name': name,
                    'emis_number': emis if emis else 'N/A'
                }
                unresolved_count += 1

            if school_info:
                resolved_count += 1
                schools_list.append(school_info)

        # Sort schools by name
        schools_list.sort(key=lambda s: s['name'])

        result.append({
            'name': aeo.get('name', 'N/A'),
            'phone': aeo.get('phone_number', 'N/A'),
            'markaz': aeo.get('markaz_name', 'N/A'),
            'schools': schools_list
        })

    print(f"Resolved from DB: {resolved_count - unresolved_count}, Used reference string as-is: {unresolved_count}")

    cur.close()
    return result


def create_excel_report(aeo_data, output_path):
    """Create a formatted Excel report of AEOs and their assigned schools."""
    wb = Workbook()
    ws = wb.active
    ws.title = "AEO Schools Report"

    # Define styles
    header_font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='left', vertical='center')

    aeo_name_font = Font(name='Calibri', bold=True, size=12, color='1F3864')
    aeo_detail_font = Font(name='Calibri', size=11, color='333333')
    aeo_label_font = Font(name='Calibri', bold=True, size=11, color='555555')

    col_header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    col_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    data_font = Font(name='Calibri', size=11)
    data_font_alt = Font(name='Calibri', size=11)
    alt_row_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin', color='B4C6E7'),
        right=Side(style='thin', color='B4C6E7'),
        top=Side(style='thin', color='B4C6E7'),
        bottom=Side(style='thin', color='B4C6E7')
    )

    no_schools_font = Font(name='Calibri', size=11, italic=True, color='999999')

    # Title row
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    title_cell = ws.cell(row=row, column=1, value="AEO Assigned Schools Report")
    title_cell.font = Font(name='Calibri', bold=True, size=16, color='1F3864')
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 30

    # Subtitle with generation date
    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    subtitle_cell = ws.cell(row=row, column=1,
                            value=f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}")
    subtitle_cell.font = Font(name='Calibri', size=10, color='666666')

    # Summary row
    row = 3
    total_aeos = len(aeo_data)
    total_schools = sum(len(a['schools']) for a in aeo_data)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    summary_cell = ws.cell(row=row, column=1,
                           value=f"Total AEOs: {total_aeos}  |  Total School Assignments: {total_schools}")
    summary_cell.font = Font(name='Calibri', size=10, color='666666', bold=True)

    row = 4  # blank separator row

    # Write each AEO section
    for aeo_idx, aeo in enumerate(aeo_data):
        row += 1

        # AEO Name header row (merged across columns)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        name_cell = ws.cell(row=row, column=1, value=f"AEO: {aeo['name']}")
        name_cell.font = aeo_name_font
        name_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 22

        # AEO Phone
        row += 1
        label_cell = ws.cell(row=row, column=1, value="Phone:")
        label_cell.font = aeo_label_font
        value_cell = ws.cell(row=row, column=2, value=aeo['phone'])
        value_cell.font = aeo_detail_font

        # AEO Markaz
        row += 1
        label_cell = ws.cell(row=row, column=1, value="Markaz:")
        label_cell.font = aeo_label_font
        value_cell = ws.cell(row=row, column=2, value=aeo['markaz'] if aeo['markaz'] else 'N/A')
        value_cell.font = aeo_detail_font

        # School count
        row += 1
        label_cell = ws.cell(row=row, column=1, value="Assigned Schools:")
        label_cell.font = aeo_label_font
        value_cell = ws.cell(row=row, column=2, value=str(len(aeo['schools'])))
        value_cell.font = aeo_detail_font

        # Column headers for schools
        row += 1
        headers = ['#', 'School Name', 'EMIS Number']
        for col_idx, header_text in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header_text)
            cell.font = col_header_font
            cell.fill = col_header_fill
            cell.alignment = Alignment(horizontal='left' if col_idx > 1 else 'center', vertical='center')
            cell.border = thin_border
        ws.row_dimensions[row].height = 20

        # School data rows
        if aeo['schools']:
            for school_idx, school in enumerate(aeo['schools'], 1):
                row += 1
                # Serial number
                num_cell = ws.cell(row=row, column=1, value=school_idx)
                num_cell.font = data_font
                num_cell.alignment = Alignment(horizontal='center')
                num_cell.border = thin_border

                # School name
                name_cell = ws.cell(row=row, column=2, value=school['name'])
                name_cell.font = data_font
                name_cell.border = thin_border

                # EMIS number
                emis_cell = ws.cell(row=row, column=3, value=school['emis_number'])
                emis_cell.font = data_font
                emis_cell.border = thin_border

                # Alternate row coloring
                if school_idx % 2 == 0:
                    num_cell.fill = alt_row_fill
                    name_cell.fill = alt_row_fill
                    emis_cell.fill = alt_row_fill
        else:
            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            no_school_cell = ws.cell(row=row, column=1, value="No schools assigned")
            no_school_cell.font = no_schools_font
            no_school_cell.alignment = Alignment(horizontal='center')

        # Add blank separator row between AEO sections
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 20

    # Freeze panes (freeze below the title area)
    ws.freeze_panes = 'A5'

    # Save
    wb.save(output_path)
    print(f"\nExcel report saved to: {output_path}")
    print(f"Total AEOs: {total_aeos}")
    print(f"Total school assignments: {total_schools}")


def main():
    db_url = get_database_url()
    print("Connecting to database...")

    conn = psycopg2.connect(db_url)
    conn.set_session(readonly=True)
    print("Connected (read-only mode)\n")

    try:
        aeo_data = query_aeo_data(conn)

        output_path = '/home/taleemabad/taleemabad/SchoolCC/AEO_Schools_Report.xlsx'
        create_excel_report(aeo_data, output_path)

        # Print summary to console
        print("\n--- Summary ---")
        for aeo in aeo_data:
            school_count = len(aeo['schools'])
            print(f"  {aeo['name']} ({aeo['phone']}) - Markaz: {aeo['markaz']} - {school_count} schools")
    finally:
        conn.close()
        print("\nDatabase connection closed.")


if __name__ == '__main__':
    main()
